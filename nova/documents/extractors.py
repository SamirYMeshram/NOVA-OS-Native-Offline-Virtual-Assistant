from __future__ import annotations

from pathlib import Path
import csv, json
from nova.config import NovaConfig

TEXT_SUFFIXES = {".txt", ".md", ".py", ".json", ".csv", ".log", ".toml", ".yaml", ".yml", ".ini"}


def extract_text(path: str | Path, config: NovaConfig | None = None) -> str:
    p = Path(path).expanduser().resolve(strict=False)
    max_bytes = config.safety.max_file_read_bytes if config else 2_000_000
    suffix = p.suffix.lower()
    if suffix in {".txt", ".md", ".py", ".log", ".toml", ".yaml", ".yml", ".ini"}:
        return p.read_text(encoding="utf-8", errors="ignore")[:max_bytes]
    if suffix == ".json":
        raw = p.read_text(encoding="utf-8", errors="ignore")[:max_bytes]
        try:
            return json.dumps(json.loads(raw), indent=2, ensure_ascii=False)
        except Exception:
            return raw
    if suffix == ".csv":
        with p.open("r", encoding="utf-8", errors="ignore", newline="") as f:
            sample = f.read(max_bytes)
        return sample
    if suffix == ".pdf":
        try:
            from pypdf import PdfReader  # type: ignore
            reader = PdfReader(str(p))
            return "\n\n".join(page.extract_text() or "" for page in reader.pages)[:max_bytes]
        except Exception as exc:
            return f"[PDF extraction unavailable: install pypdf. Error: {exc}]"
    if suffix == ".docx":
        try:
            import docx  # type: ignore
            doc = docx.Document(str(p))
            return "\n".join(par.text for par in doc.paragraphs)[:max_bytes]
        except Exception as exc:
            return f"[DOCX extraction unavailable: install python-docx. Error: {exc}]"
    if suffix in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(p, read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in wb.worksheets:
                lines.append(f"# Sheet: {sheet.title}")
                for row in sheet.iter_rows(values_only=True):
                    lines.append(", ".join("" if c is None else str(c) for c in row))
                    if sum(len(x) for x in lines) > max_bytes: break
            return "\n".join(lines)[:max_bytes]
        except Exception as exc:
            return f"[Excel extraction unavailable: install openpyxl. Error: {exc}]"
    return f"[Unsupported file type: {suffix}]"


def supported(path: str | Path) -> bool:
    return Path(path).suffix.lower() in TEXT_SUFFIXES | {".pdf", ".docx", ".xlsx", ".xlsm"}
